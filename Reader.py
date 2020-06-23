from openpyxl import load_workbook
from openpyxl import Workbook
import sqlite3
# SQLite DB 연결
conn = sqlite3.connect("Yesan.db")
cur = conn.cursor()
global Yesan
global Pum
global Spend
Yesan = '''CREATE TABLE "Yesan" (
	"Team"	TEXT,
	"Busy"	TEXT,
	"Budget"	INTEGER,
	PRIMARY KEY("Team","Busy")
)'''
Pum = '''CREATE TABLE "Pum" (
	"Numb"	TEXT,
	"Team"	TEXT,
	"Busy"	TEXT,
	"Desc"	TEXT,
	"Budget"	INTEGER,
	PRIMARY KEY("Numb"),
	FOREIGN KEY("Busy") REFERENCES "Yesan"("Busy"),
	FOREIGN KEY("Team") REFERENCES "Yesan"("Team")
)'''
Spend = '''CREATE TABLE "Spend" (
	"Numb"	TEXT,
	"Date"	TEXT,
	"Pumnumb"	TEXT,
	"Subject"	TEXT,
	"Desc"	TEXT,
	"Prov"	INTEGER,
	"Tax"	INTEGER,
	FOREIGN KEY("Pumnumb") REFERENCES "Pum"("Numb"),
	PRIMARY KEY("Numb")
)'''
global printTable
def printTable(Table,conn):
    print("Print "+Table)
    #conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    sql = "select * from "+Table
    rows = cur.execute(sql)
    for row in rows:
        print(row)
    conn.close()
global makeTable
def makeTable(sql,conn):
    #conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    cur.execute(sql)
    conn.commit()
    #conn.close()

def clearTable(Table,conn):
    #conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    sql = "delete from "+Table
    result = cur.execute(sql)
    conn.commit()
    #conn.close()

def dropTable(Table,conn):
    #conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    sql = "drop table " + Table
    result = cur.execute(sql)
    conn.commit()
    #conn.close()
def insertYesan(conn):
    load_wb = load_workbook("C:/예제T.xlsx", data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['예산']
    #conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    insert = "insert into Yesan(Team, Busy, Budget) values (?, ?, ?)"
    get_cells = load_ws['B5':'D100']
    flag = False
    for row in get_cells:
        if flag:
            break
        tmp = []
        for cell in row:
            if cell.value is None:
                flag = True
                break
            else:
                tmp.append(cell.value)
               #print(cell.value)
        if flag:
            break
        #print(tmp)
        cur.execute(insert, (tmp[0], tmp[1], tmp[2]))
        conn.commit()
    #conn.close()
    print("insert Yesan Finished")

def insertPum(conn):
    load_wb = load_workbook("C:/Users/Heun/Desktop/예제T.xlsx", data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['품의']
    #conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    insert = "insert into Pum(Team, Busy, Desc, Numb, Budget) values (?, ?, ?, ?, ?)"
    get_cells = load_ws['B5':'F200']
    flag = False
    for row in get_cells:
        if flag:
            break
        tmp = []
        for cell in row:
            if cell.value is None:
                flag = True
                break
            else:
                tmp.append(cell.value)
                #print(cell.value)
        if flag:
            break
        #print(tmp)
        cur.execute(insert, (tmp[0], tmp[1], tmp[2], tmp[3], tmp[4]))
        conn.commit()
    #conn.close()
    print("insert Pum Finished")
def insertspend(conn):
    load_wb = load_workbook("C:/Users/Heun/Desktop/예제T.xlsx", data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['지출입력']
    #conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    insert = "insert into Spend(Date, Pumnumb, Numb, Subject, Desc, Prov, Tax) values (?, ?, ?, ?, ?, ?, ?)"
    get_cells = load_ws['B6':'J300']
    flag = False
    for row in get_cells:
        if flag:
            break
        tmp = []
        chk = True
        for cell in row:
            if chk:
                if cell.value is None:
                    flag = True
                    break
                else:
                    chk = False
                    tmp.append(cell.value)
                    #print(cell.value)
            else:
                tmp.append(cell.value)
                #print(cell.value)
        if flag:
            break
        #print(tmp)
        date = tmp[0].strftime('%Y-%m-%d')
        cur.execute(insert, (date, tmp[2], tmp[3], tmp[4], tmp[5],tmp[6],tmp[7]))
        conn.commit()
    #conn.close()
    print("insert Spend Finished")
def init(conn):
    try:
        makeTable(Yesan,conn)
    except:
        dropTable("Yesan",conn)
        makeTable(Yesan,conn)
    try:
        makeTable(Pum,conn)
    except:
        dropTable("Pum",conn)
        makeTable(Pum,conn)
    try:
        makeTable(Spend,conn)
    except:
        dropTable("Spend",conn)
        makeTable(Spend,conn)
    insertYesan(conn)
    insertPum(conn)
    insertspend(conn)
    #conn.close()

conn.close()