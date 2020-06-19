from openpyxl import load_workbook
from openpyxl import Workbook
import sqlite3

# 셀 주소로 값 출력
#print(load_ws['C7'].value)
# 셀 좌표로 값 출력
#print(load_ws.cell(7, 3).value)

# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.

def insertYesan():
    load_wb = load_workbook("C:/Users/Heun/Desktop/예제T.xlsx", data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['예산']
    conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    insert = "insert into Yesan(Team, Busy, Budget) values (?, ?, ?)"
    get_cells = load_ws['B5':'D100']
    flag = False
    for row in get_cells:
        if flag:
            break
        tmp = []
        for cell in row:
            if cell.value is None:
                flag = True
                break
            else:
                tmp.append(cell.value)
                print(cell.value)
        if flag:
            break
        print(tmp)
        cur.execute(insert, (tmp[0], tmp[1], tmp[2]))
        conn.commit()
    conn.close()

def insertPum():
    load_wb = load_workbook("C:/Users/Heun/Desktop/예제T.xlsx", data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['품의']
    conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    insert = "insert into Pum(Team, Busy, Desc, Numb, Budget) values (?, ?, ?, ?, ?)"
    get_cells = load_ws['B5':'F200']
    flag = False
    for row in get_cells:
        if flag:
            break
        tmp = []
        for cell in row:
            if cell.value is None:
                flag = True
                break
            else:
                tmp.append(cell.value)
                #print(cell.value)
        if flag:
            break
        print(tmp)
        cur.execute(insert, (tmp[0], tmp[1], tmp[2], tmp[3], tmp[4]))
        conn.commit()
    conn.close()

def insertspend():
    load_wb = load_workbook("C:/Users/Heun/Desktop/예제T.xlsx", data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['지출입력']
    conn = sqlite3.connect("Yesan.db")
    cur = conn.cursor()
    insert = "insert into Spend(Date, Pumnumb, Numb, Subject, Desc, Prov, Tax) values (?, ?, ?, ?, ?, ?, ?)"
    get_cells = load_ws['B6':'J300']
    flag = False
    for row in get_cells:
        if flag:
            break
        tmp = []
        chk = True
        for cell in row:
            if chk:
                if cell.value is None:
                    flag = True
                    break
                else:
                    chk = False
                    tmp.append(cell.value)
                    #print(cell.value)
            else:
                tmp.append(cell.value)
                #print(cell.value)
        if flag:
            break
        print(tmp)
        date = tmp[0].strftime('%Y-%m-%d')
        cur.execute(insert, (date, tmp[2], tmp[3], tmp[4], tmp[5],tmp[6],tmp[7]))
        conn.commit()
    conn.close()
    print("insert Finished")
#insertPum()
insertspend()

# print(load_ws.append(range(22)))
# load_ws.append(['이경헌'])
# load_ws.title = "다음 이름"
# load_wb.save("테스트2.xlsx")